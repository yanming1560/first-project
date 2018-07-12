from openpyxl import load_workbook as lw
from openpyxl.drawing.image import Image

f=lw('C:\\Users\cnstyan9\Desktop\\test.xlsx')   #打开excel
w1=f['Sheet1']      #指定worksheet
print(w1['c12'].value)  #输出worksheet中指定单元格的值
f.create_sheet('shttt')   #新建worksheet
w2=f['shttt']
w2['a1']='hhhahah'      #指定单元格赋值
img=Image('3.jpg')      #指定图片
w2.add_image(img,'a2')      #指定位置插入图片
print(w2.cell(row=1,column=1).value)
w2.cell(row=1,column=2).value=10
f.save('hha.xlsx')      #文件要保存
